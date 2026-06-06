"""
告警管理API - 重构版
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_
from typing import Optional, List
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
from app.core.database import get_db
from app.core.timezone import get_timezone
from app.models.user import User
from app.models.server import Server
from app.models.alert import (
    AlertContact, AlertGroup, AlertGroupMember, AlertRule, 
    AlertRuleNotification, Alert, AlertSilence, AlertStatus
)
from app.schemas.alert import (
    AlertContactCreate, AlertContactUpdate, AlertContactResponse,
    AlertGroupCreate, AlertGroupUpdate, AlertGroupResponse,
    AlertRuleCreate, AlertRuleUpdate, AlertRuleResponse,
    AlertResponse, AlertListResponse,
    AlertSilenceCreate, AlertSilenceResponse
)
from app.utils.dependencies import get_current_user
import logging

logger = logging.getLogger(__name__)

router = APIRouter()


def _load_rules_map(db: Session, alerts: List[Alert]) -> dict:
    rule_ids = {a.rule_id for a in alerts if a.rule_id}
    if not rule_ids:
        return {}
    return {
        r.id: r for r in db.query(AlertRule).filter(AlertRule.id.in_(rule_ids)).all()
    }


def _build_alert_response(alert: Alert, rule: Optional[AlertRule] = None) -> AlertResponse:
    return AlertResponse(
        id=alert.id,
        rule_id=alert.rule_id,
        rule_name=alert.rule_name or (rule.name if rule else None),
        server_id=alert.server_id,
        metric_type=alert.metric_type,
        metric_label=alert.metric_label,
        current_value=alert.current_value,
        threshold=alert.threshold,
        operator=alert.operator or (rule.operator if rule else '>') or '>',
        severity=alert.severity.value if alert.severity else None,
        status=alert.status.value if alert.status else None,
        fired_at=alert.fired_at,
        last_occurred_at=alert.last_occurred_at,
        resolved_at=alert.resolved_at,
    )


# ===== 告警人管理 =====
@router.get("/contacts", response_model=List[AlertContactResponse])
def get_contacts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取告警人列表"""
    contacts = db.query(AlertContact).all()
    return [AlertContactResponse(**c.to_dict()) for c in contacts]


@router.post("/contacts", response_model=AlertContactResponse, status_code=status.HTTP_201_CREATED)
def create_contact(
    data: AlertContactCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建告警人"""
    contact = AlertContact(**data.model_dump())
    db.add(contact)
    db.commit()
    db.refresh(contact)
    logger.info(f"User {current_user.username} created contact: {contact.name}")
    return AlertContactResponse(**contact.to_dict())


@router.put("/contacts/{contact_id}", response_model=AlertContactResponse)
def update_contact(
    contact_id: int,
    data: AlertContactUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新告警人"""
    contact = db.query(AlertContact).filter(AlertContact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(contact, field, value)
    
    db.commit()
    db.refresh(contact)
    return AlertContactResponse(**contact.to_dict())


@router.delete("/contacts/{contact_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_contact(
    contact_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除告警人"""
    contact = db.query(AlertContact).filter(AlertContact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    
    db.delete(contact)
    db.commit()
    return None


# ===== 告警组管理 =====
@router.get("/groups", response_model=List[AlertGroupResponse])
def get_groups(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取告警组列表"""
    groups = db.query(AlertGroup).all()
    result = []
    for group in groups:
        group_dict = group.to_dict()
        # 获取组成员
        members = db.query(AlertGroupMember).filter(AlertGroupMember.group_id == group.id).all()
        contact_ids = [m.contact_id for m in members]
        contacts = db.query(AlertContact).filter(AlertContact.id.in_(contact_ids)).all() if contact_ids else []
        group_dict["contacts"] = [AlertContactResponse(**c.to_dict()) for c in contacts]
        result.append(AlertGroupResponse(**group_dict))
    return result


@router.post("/groups", response_model=AlertGroupResponse, status_code=status.HTTP_201_CREATED)
def create_group(
    data: AlertGroupCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建告警组"""
    group = AlertGroup(name=data.name, description=data.description)
    db.add(group)
    db.commit()
    db.refresh(group)
    
    # 添加成员
    for contact_id in data.contact_ids:
        member = AlertGroupMember(group_id=group.id, contact_id=contact_id)
        db.add(member)
    db.commit()
    
    logger.info(f"User {current_user.username} created group: {group.name}")
    
    # 返回完整数据
    contacts = db.query(AlertContact).filter(AlertContact.id.in_(data.contact_ids)).all() if data.contact_ids else []
    group_dict = group.to_dict()
    group_dict["contacts"] = [AlertContactResponse(**c.to_dict()) for c in contacts]
    return AlertGroupResponse(**group_dict)


@router.put("/groups/{group_id}", response_model=AlertGroupResponse)
def update_group(
    group_id: int,
    data: AlertGroupUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新告警组"""
    group = db.query(AlertGroup).filter(AlertGroup.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    
    if data.name:
        group.name = data.name
    if data.description is not None:
        group.description = data.description
    
    # 更新成员
    if data.contact_ids is not None:
        db.query(AlertGroupMember).filter(AlertGroupMember.group_id == group_id).delete()
        for contact_id in data.contact_ids:
            member = AlertGroupMember(group_id=group_id, contact_id=contact_id)
            db.add(member)
    
    db.commit()
    db.refresh(group)
    
    # 返回完整数据
    members = db.query(AlertGroupMember).filter(AlertGroupMember.group_id == group_id).all()
    contact_ids = [m.contact_id for m in members]
    contacts = db.query(AlertContact).filter(AlertContact.id.in_(contact_ids)).all() if contact_ids else []
    group_dict = group.to_dict()
    group_dict["contacts"] = [AlertContactResponse(**c.to_dict()) for c in contacts]
    return AlertGroupResponse(**group_dict)


@router.delete("/groups/{group_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_group(
    group_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除告警组"""
    group = db.query(AlertGroup).filter(AlertGroup.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    
    db.delete(group)
    db.commit()
    return None


# ===== 告警规则管理 =====
@router.get("/rules", response_model=List[AlertRuleResponse])
def get_rules(
    server_id: Optional[int] = Query(None),
    enabled: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取告警规则列表"""
    query = db.query(AlertRule)
    if server_id:
        query = query.filter(AlertRule.server_id == server_id)
    if enabled is not None:
        query = query.filter(AlertRule.enabled == enabled)
    
    rules = query.all()
    result = []
    for rule in rules:
        rule_dict = rule.to_dict()
        # 获取通知配置
        notifications = db.query(AlertRuleNotification).filter(AlertRuleNotification.rule_id == rule.id).all()
        contact_ids = [n.contact_id for n in notifications if n.contact_id]
        group_ids = [n.group_id for n in notifications if n.group_id]
        
        contacts = db.query(AlertContact).filter(AlertContact.id.in_(contact_ids)).all() if contact_ids else []
        groups = db.query(AlertGroup).filter(AlertGroup.id.in_(group_ids)).all() if group_ids else []
        
        rule_dict["contacts"] = [AlertContactResponse(**c.to_dict()) for c in contacts]
        rule_dict["groups"] = [AlertGroupResponse(**g.to_dict()) for g in groups]
        result.append(AlertRuleResponse(**rule_dict))
    
    return result


@router.post("/rules", response_model=AlertRuleResponse, status_code=status.HTTP_201_CREATED)
def create_rule(
    data: AlertRuleCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建告警规则"""
    # 检查服务器是否存在
    server = db.query(Server).filter(Server.id == data.server_id).first()
    if not server:
        raise HTTPException(status_code=404, detail="Server not found")
    
    rule = AlertRule(
        name=data.name,
        description=data.description,
        server_id=data.server_id,
        metric_type=data.metric_type,
        metric_label=data.metric_label,
        operator=data.operator,
        threshold=data.threshold,
        duration=data.duration,
        repeat_interval=data.repeat_interval,
        severity=data.severity,
        enabled=data.enabled
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    
    # 添加通知配置
    for contact_id in data.contact_ids:
        notif = AlertRuleNotification(rule_id=rule.id, contact_id=contact_id)
        db.add(notif)
    for group_id in data.group_ids:
        notif = AlertRuleNotification(rule_id=rule.id, group_id=group_id)
        db.add(notif)
    db.commit()
    
    logger.info(f"User {current_user.username} created rule: {rule.name}")
    
    # 返回完整数据
    contacts = db.query(AlertContact).filter(AlertContact.id.in_(data.contact_ids)).all() if data.contact_ids else []
    groups = db.query(AlertGroup).filter(AlertGroup.id.in_(data.group_ids)).all() if data.group_ids else []
    rule_dict = rule.to_dict()
    rule_dict["contacts"] = [AlertContactResponse(**c.to_dict()) for c in contacts]
    rule_dict["groups"] = [AlertGroupResponse(**g.to_dict()) for g in groups]
    return AlertRuleResponse(**rule_dict)


@router.put("/rules/{rule_id}", response_model=AlertRuleResponse)
def update_rule(
    rule_id: int,
    data: AlertRuleUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新告警规则"""
    rule = db.query(AlertRule).filter(AlertRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    
    update_data = data.model_dump(exclude_unset=True, exclude={"contact_ids", "group_ids"})
    for field, value in update_data.items():
        setattr(rule, field, value)

    if rule.metric_type == "host_up":
        rule.operator = "<"
        rule.threshold = 1
        rule.metric_label = None
    
    # 更新通知配置
    if data.contact_ids is not None or data.group_ids is not None:
        db.query(AlertRuleNotification).filter(AlertRuleNotification.rule_id == rule_id).delete()
        if data.contact_ids:
            for contact_id in data.contact_ids:
                notif = AlertRuleNotification(rule_id=rule_id, contact_id=contact_id)
                db.add(notif)
        if data.group_ids:
            for group_id in data.group_ids:
                notif = AlertRuleNotification(rule_id=rule_id, group_id=group_id)
                db.add(notif)
    
    db.commit()
    db.refresh(rule)
    
    # 返回完整数据
    notifications = db.query(AlertRuleNotification).filter(AlertRuleNotification.rule_id == rule_id).all()
    contact_ids = [n.contact_id for n in notifications if n.contact_id]
    group_ids = [n.group_id for n in notifications if n.group_id]
    contacts = db.query(AlertContact).filter(AlertContact.id.in_(contact_ids)).all() if contact_ids else []
    groups = db.query(AlertGroup).filter(AlertGroup.id.in_(group_ids)).all() if group_ids else []
    rule_dict = rule.to_dict()
    rule_dict["contacts"] = [AlertContactResponse(**c.to_dict()) for c in contacts]
    rule_dict["groups"] = [AlertGroupResponse(**g.to_dict()) for g in groups]
    return AlertRuleResponse(**rule_dict)


@router.delete("/rules/{rule_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_rule(
    rule_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除告警规则"""
    rule = db.query(AlertRule).filter(AlertRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    
    db.delete(rule)
    db.commit()
    return None


# ===== 当前告警 =====
@router.get("/current", response_model=AlertListResponse)
def get_current_alerts(
    server_id: Optional[int] = Query(None),
    severity: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取当前告警（FIRING 状态且未被有效屏蔽的告警）"""
    now = datetime.now()

    # 子查询：当前有效的屏蔽覆盖的 (rule_id, server_id) 组合
    active_silence_subq = (
        db.query(AlertSilence.rule_id, AlertSilence.server_id)
        .filter(
            AlertSilence.is_active == True,
            AlertSilence.expires_at > now
        )
        .subquery()
    )

    query = (
        db.query(Alert)
        .outerjoin(
            active_silence_subq,
            and_(
                Alert.rule_id == active_silence_subq.c.rule_id,
                Alert.server_id == active_silence_subq.c.server_id
            )
        )
        .filter(
            Alert.status == AlertStatus.FIRING,
            active_silence_subq.c.rule_id == None  # 没有匹配到有效屏蔽
        )
    )
    
    if server_id:
        query = query.filter(Alert.server_id == server_id)
    if severity:
        query = query.filter(Alert.severity == severity)
    
    total = query.count()
    alerts = query.order_by(Alert.fired_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    rules_map = _load_rules_map(db, alerts)
    items = [
        _build_alert_response(alert, rules_map.get(alert.rule_id))
        for alert in alerts
    ]
    
    return AlertListResponse(
        total=total,
        page=page,
        page_size=page_size,
        items=items
    )


# ===== 历史告警 =====
@router.get("/history", response_model=AlertListResponse)
def get_alert_history(
    server_id: Optional[int] = Query(None),
    start_time: Optional[datetime] = Query(None),
    end_time: Optional[datetime] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取历史告警"""
    query = db.query(Alert).filter(Alert.status == AlertStatus.RESOLVED)
    
    if server_id:
        query = query.filter(Alert.server_id == server_id)
    if start_time:
        query = query.filter(Alert.fired_at >= start_time)
    if end_time:
        query = query.filter(Alert.fired_at <= end_time)
    
    total = query.count()
    alerts = query.order_by(Alert.fired_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    rules_map = _load_rules_map(db, alerts)
    items = [
        _build_alert_response(alert, rules_map.get(alert.rule_id))
        for alert in alerts
    ]
    
    return AlertListResponse(
        total=total,
        page=page,
        page_size=page_size,
        items=items
    )


@router.get("/history/export")
def export_alert_history(
    server_id: Optional[int] = Query(None),
    start_time: Optional[datetime] = Query(None),
    end_time: Optional[datetime] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出历史告警为XLSX"""
    query = db.query(Alert).filter(Alert.status == AlertStatus.RESOLVED)
    
    if server_id:
        query = query.filter(Alert.server_id == server_id)
    if start_time:
        query = query.filter(Alert.fired_at >= start_time)
    if end_time:
        query = query.filter(Alert.fired_at <= end_time)
    
    alerts = query.order_by(Alert.fired_at.desc()).all()
    
    # 获取服务器信息映射
    servers = {s.id: s for s in db.query(Server).all()}
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "历史告警"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    headers = ["规则名称", "服务器", "指标类型", "当前值", "阈值", "级别", "触发时间", "最后发生时间", "恢复时间"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 级别映射
    severity_map = {
        'level1': '1级告警',
        'level2': '2级告警',
        'level3': '3级告警',
        'level4': '4级告警'
    }

    # 获取告警规则 operator 映射（rule_id -> operator）
    rule_ids = list({a.rule_id for a in alerts if a.rule_id})
    rules_map = {}
    if rule_ids:
        rules_map = {r.id: r for r in db.query(AlertRule).filter(AlertRule.id.in_(rule_ids)).all()}

    def format_metric_type(metric_type: str, metric_label: str) -> str:
        """将指标类型格式化为可读字符串"""
        if metric_type == 'cpu':
            return 'CPU使用率'
        elif metric_type == 'memory':
            return '内存使用率'
        elif metric_type == 'disk_mount':
            label = metric_label or ''
            return f'磁盘使用率:{label}'
        elif metric_type == 'tcp_conn':
            label = metric_label or ''
            tcp_type_map = {"CurrEstab": "已建立", "AllEstab": "所有"}
            return f'TCP连接数:{tcp_type_map.get(label, label)}'
        elif metric_type == 'host_up':
            return '主机存活'
        return metric_type

    def format_value(value: float, metric_type: str) -> str:
        """格式化监控值"""
        if metric_type == 'tcp_conn':
            return f"{int(value)}个"
        if metric_type == 'host_up':
            return "离线" if value < 1 else "在线"
        return f"{value:.2f}%"

    def format_threshold(operator: str, threshold: float, metric_type: str) -> str:
        """格式化阈值，带上操作符和单位"""
        if metric_type == 'tcp_conn':
            return f"{operator} {int(threshold)}个"
        if metric_type == 'host_up':
            return f"{operator} {int(threshold)}"
        return f"{operator} {threshold:.2f}%"

    # 写入数据
    for row_num, alert in enumerate(alerts, 2):
        server = servers.get(alert.server_id)
        server_name = f"{server.hostname} ({server.ip_address})" if server else "未知"
        rule = rules_map.get(alert.rule_id) if alert.rule_id else None
        rule_name = alert.rule_name or (rule.name if rule else "未知")
        operator = alert.operator or (rule.operator if rule else '>')

        ws.cell(row=row_num, column=1, value=rule_name)
        ws.cell(row=row_num, column=2, value=server_name)
        ws.cell(row=row_num, column=3, value=format_metric_type(alert.metric_type, alert.metric_label))
        ws.cell(row=row_num, column=4, value=format_value(alert.current_value, alert.metric_type))
        ws.cell(row=row_num, column=5, value=format_threshold(operator, alert.threshold, alert.metric_type))
        ws.cell(row=row_num, column=6, value=severity_map.get(str(alert.severity.value if hasattr(alert.severity, 'value') else alert.severity), str(alert.severity)))
        ws.cell(row=row_num, column=7, value=alert.fired_at.strftime("%Y-%m-%d %H:%M:%S") if alert.fired_at else "")
        ws.cell(row=row_num, column=8, value=alert.last_occurred_at.strftime("%Y-%m-%d %H:%M:%S") if alert.last_occurred_at else "")
        ws.cell(row=row_num, column=9, value=alert.resolved_at.strftime("%Y-%m-%d %H:%M:%S") if alert.resolved_at else "")
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 生成文件名
    filename = f"alert_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 返回Excel文件
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ===== 告警屏蔽 =====
@router.get("/silences")
def get_silences(
    is_active: Optional[bool] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取告警屏蔽列表"""
    query = db.query(AlertSilence)
    
    if is_active is not None:
        query = query.filter(AlertSilence.is_active == is_active)
    
    total = query.count()
    silences = query.order_by(AlertSilence.silenced_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    # 增强返回数据，添加关联信息
    items = []
    for silence in silences:
        silence_dict = silence.to_dict()
        
        # 添加规则名称
        if silence.rule_id:
            rule = db.query(AlertRule).filter(AlertRule.id == silence.rule_id).first()
            silence_dict['rule_name'] = rule.name if rule else None
        else:
            silence_dict['rule_name'] = None
        
        # 添加服务器名称
        if silence.server_id:
            server = db.query(Server).filter(Server.id == silence.server_id).first()
            silence_dict['server_name'] = f"{server.hostname} ({server.ip_address})" if server else None
        else:
            silence_dict['server_name'] = None
        
        items.append(silence_dict)
    
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": items
    }


@router.post("/silences", response_model=AlertSilenceResponse, status_code=status.HTTP_201_CREATED)
def create_silence(
    data: AlertSilenceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建告警屏蔽"""
    
    # 计算过期时间
    expires_at = data.expires_at
    if expires_at is None and data.duration_minutes:
        # 使用应用时区计算过期时间
        expires_at = datetime.now(get_timezone()) + timedelta(minutes=data.duration_minutes)
    elif expires_at:
        # 如果前端传来的时间没有时区信息，假定为应用时区
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=get_timezone())
    
    silence = AlertSilence(
        alert_id=data.alert_id,
        rule_id=data.rule_id,
        server_id=data.server_id,
        reason=data.reason,
        expires_at=expires_at,
        is_active=True
    )
    db.add(silence)
    db.commit()
    db.refresh(silence)
    
    logger.info(f"User {current_user.username} created silence for rule_id={data.rule_id}, server_id={data.server_id}, expires_at={expires_at}")
    return AlertSilenceResponse(**silence.to_dict())


@router.delete("/silences/{silence_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_silence(
    silence_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """取消告警屏蔽"""
    silence = db.query(AlertSilence).filter(AlertSilence.id == silence_id).first()
    if not silence:
        raise HTTPException(status_code=404, detail="Silence not found")
    
    # 将屏蔽标记为不活跃
    silence.is_active = False
    
    logger.info(f"User {current_user.username} cancelled silence {silence_id}")
    db.commit()
    return None
