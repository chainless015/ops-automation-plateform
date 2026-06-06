"""
脚本执行API
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
from app.core.database import get_db
from app.models.user import User
from app.models.script import Script
from app.models.server import Server
from app.models.execution import ExecutionHistory
from app.schemas.execution import ExecutionCreate, ExecutionListItem, ExecutionDetail, ExecutionResponse
from app.utils.dependencies import get_current_user
from app.services.ssh_service import execute_script_via_ssh, SSHConnectionError, SSHExecutionError
import logging

logger = logging.getLogger(__name__)

router = APIRouter()

# 上海时区
SHANGHAI_TZ = ZoneInfo("Asia/Shanghai")


def execute_script_task(execution_id: int, script_id: int, server_id: int):
    """
    后台任务：执行脚本
    
    Args:
        execution_id: 执行记录ID
        script_id: 脚本ID
        server_id: 服务器ID
    """
    # 创建新的数据库会话
    from app.core.database import SessionLocal
    db = SessionLocal()
    
    try:
        # 获取脚本和服务器信息
        script = db.query(Script).filter(Script.id == script_id).first()
        server = db.query(Server).filter(Server.id == server_id).first()
        
        if not script or not server:
            # 更新执行状态为失败
            execution = db.query(ExecutionHistory).filter(ExecutionHistory.id == execution_id).first()
            if execution:
                execution.status = "failed"
                execution.stderr = "Script or server not found"
                db.commit()
            return
        
        # 执行脚本
        return_code, stdout, stderr, duration = execute_script_via_ssh(
            server=server,
            script_content=script.content,
            timeout=300
        )
        
        # 更新执行记录
        execution = db.query(ExecutionHistory).filter(ExecutionHistory.id == execution_id).first()
        if execution:
            execution.status = "success" if return_code == 0 else "failed"
            execution.return_code = return_code
            execution.stdout = stdout
            execution.stderr = stderr
            execution.duration_seconds = duration
            execution.finished_at = datetime.now(SHANGHAI_TZ)
            db.commit()
            
            logger.info(f"Execution {execution_id} completed: status={execution.status}, return_code={return_code}")
        
    except SSHConnectionError as e:
        logger.error(f"SSH connection error for execution {execution_id}: {str(e)}")
        execution = db.query(ExecutionHistory).filter(ExecutionHistory.id == execution_id).first()
        if execution:
            execution.status = "failed"
            execution.stderr = str(e)
            execution.finished_at = datetime.now(SHANGHAI_TZ)
            db.commit()
    except SSHExecutionError as e:
        logger.error(f"SSH execution error for execution {execution_id}: {str(e)}")
        execution = db.query(ExecutionHistory).filter(ExecutionHistory.id == execution_id).first()
        if execution:
            execution.status = "failed"
            execution.stderr = str(e)
            execution.finished_at = datetime.now(SHANGHAI_TZ)
            db.commit()
    except Exception as e:
        logger.error(f"Unexpected error for execution {execution_id}: {str(e)}")
        execution = db.query(ExecutionHistory).filter(ExecutionHistory.id == execution_id).first()
        if execution:
            execution.status = "failed"
            execution.stderr = f"Unexpected error: {str(e)}"
            execution.finished_at = datetime.now(SHANGHAI_TZ)
            db.commit()
    finally:
        db.close()


@router.post("", response_model=ExecutionResponse, status_code=status.HTTP_202_ACCEPTED)
def execute_script(
    execution_data: ExecutionCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    执行脚本（异步）
    
    Args:
        execution_data: 执行数据（包含script_id和server_id）
        background_tasks: 后台任务
        db: 数据库会话
        current_user: 当前用户
    
    Returns:
        执行响应
    
    Raises:
        HTTPException: 脚本或服务器不存在时抛出404错误
    """
    # 验证脚本存在
    script = db.query(Script).filter(Script.id == execution_data.script_id).first()
    if not script:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Script not found"
        )
    
    # 验证服务器存在
    server = db.query(Server).filter(Server.id == execution_data.server_id).first()
    if not server:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Server not found"
        )
    
    # 创建执行记录
    execution = ExecutionHistory(
        script_id=execution_data.script_id,
        server_id=execution_data.server_id,
        status="running",
        execution_type="manual"
    )
    
    db.add(execution)
    db.commit()
    db.refresh(execution)
    
    # 添加后台任务
    background_tasks.add_task(
        execute_script_task,
        execution.id,
        execution_data.script_id,
        execution_data.server_id
    )
    
    logger.info(f"User {current_user.username} started execution {execution.id} for script {script.name} on server {server.ip_address}")
    
    return ExecutionResponse(
        execution_id=execution.id,
        status="running",
        message="Script execution started"
    )


@router.get("")
def get_executions(
    server_id: Optional[int] = Query(None, description="服务器ID筛选"),
    script_id: Optional[int] = Query(None, description="脚本ID筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    execution_type: Optional[str] = Query(None, description="执行类型筛选"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取执行历史列表
    
    Args:
        server_id: 服务器ID筛选
        script_id: 脚本ID筛选
        status: 状态筛选
        execution_type: 执行类型筛选（manual/scheduled）
        start_date: 开始日期
        end_date: 结束日期
        page: 页码
        page_size: 每页数量
        db: 数据库会话
        current_user: 当前用户
    
    Returns:
        执行历史列表（带分页信息）
    """
    query = db.query(ExecutionHistory)
    
    # 筛选条件
    if server_id:
        query = query.filter(ExecutionHistory.server_id == server_id)
    if script_id:
        query = query.filter(ExecutionHistory.script_id == script_id)
    if status:
        query = query.filter(ExecutionHistory.status == status)
    if execution_type:
        query = query.filter(ExecutionHistory.execution_type == execution_type)
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(ExecutionHistory.executed_at >= start_dt)
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(ExecutionHistory.executed_at < end_dt)
        except ValueError:
            pass
    
    # 获取总数
    total = query.count()
    
    # 排序和分页
    executions = query.order_by(ExecutionHistory.executed_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    # 获取脚本和服务器信息映射
    scripts = {s.id: s for s in db.query(Script).all()}
    servers = {s.id: s for s in db.query(Server).all()}
    
    # 构建响应，包含脚本和服务器名称
    items = []
    for execution in executions:
        data = execution.to_dict(include_output=False)
        script = scripts.get(execution.script_id)
        server = servers.get(execution.server_id)
        data['script_name'] = script.name if script else "未知"
        data['server_hostname'] = server.hostname if server else "未知"
        data['server_ip'] = server.ip_address if server else "未知"
        items.append(ExecutionListItem(**data))
    
    return {
        "items": items,
        "total": total
    }


@router.get("/export")
def export_executions(
    ids: Optional[str] = Query(None, description="执行记录ID列表，逗号分隔"),
    server_id: Optional[int] = Query(None, description="服务器ID筛选"),
    execution_type: Optional[str] = Query(None, description="执行类型筛选"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    导出执行历史为XLSX
    
    Args:
        ids: 执行记录ID列表（导出选中时使用）
        server_id: 服务器ID筛选
        execution_type: 执行类型筛选
        start_date: 开始日期
        end_date: 结束日期
        db: 数据库会话
        current_user: 当前用户
    
    Returns:
        XLSX文件流
    """
    query = db.query(ExecutionHistory)
    
    # 如果提供了ID列表，只导出这些记录
    if ids:
        id_list = [int(id.strip()) for id in ids.split(',') if id.strip()]
        query = query.filter(ExecutionHistory.id.in_(id_list))
    else:
        # 否则使用筛选条件
        if server_id:
            query = query.filter(ExecutionHistory.server_id == server_id)
        if execution_type:
            query = query.filter(ExecutionHistory.execution_type == execution_type)
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                query = query.filter(ExecutionHistory.executed_at >= start_dt)
            except ValueError:
                pass
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
                query = query.filter(ExecutionHistory.executed_at < end_dt)
            except ValueError:
                pass
    
    # 获取所有执行记录
    executions = query.order_by(ExecutionHistory.executed_at.desc()).all()
    
    # 获取脚本和服务器信息映射
    scripts = {s.id: s for s in db.query(Script).all()}
    servers = {s.id: s for s in db.query(Server).all()}
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "执行历史"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    headers = ["脚本名称", "服务器", "执行类型", "状态", "退出码", "执行时长(秒)", "开始时间", "结束时间", "标准输出", "错误输出"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 写入数据
    for row_num, execution in enumerate(executions, 2):
        script = scripts.get(execution.script_id)
        script_name = script.name if script else "未知脚本"
        
        server = servers.get(execution.server_id)
        server_name = f"{server.hostname} ({server.ip_address})" if server else "未知服务器"
        
        execution_type_text = "手动执行" if execution.execution_type == "manual" else "定时任务"
        
        ws.cell(row=row_num, column=1, value=script_name)
        ws.cell(row=row_num, column=2, value=server_name)
        ws.cell(row=row_num, column=3, value=execution_type_text)
        ws.cell(row=row_num, column=4, value=execution.status)
        ws.cell(row=row_num, column=5, value=execution.return_code if execution.return_code is not None else "")
        ws.cell(row=row_num, column=6, value=f"{execution.duration_seconds:.2f}" if execution.duration_seconds is not None else "")
        ws.cell(row=row_num, column=7, value=execution.executed_at.strftime("%Y-%m-%d %H:%M:%S") if execution.executed_at else "")
        ws.cell(row=row_num, column=8, value=execution.finished_at.strftime("%Y-%m-%d %H:%M:%S") if execution.finished_at else "")
        ws.cell(row=row_num, column=9, value=execution.stdout or "")
        ws.cell(row=row_num, column=10, value=execution.stderr or "")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15  # 脚本名称
    ws.column_dimensions['B'].width = 30  # 服务器
    ws.column_dimensions['C'].width = 10  # 执行类型
    ws.column_dimensions['D'].width = 10  # 状态
    ws.column_dimensions['E'].width = 10  # 退出码
    ws.column_dimensions['F'].width = 15  # 执行时长
    ws.column_dimensions['G'].width = 20  # 开始时间
    ws.column_dimensions['H'].width = 20  # 结束时间
    ws.column_dimensions['I'].width = 25  # 标准输出
    ws.column_dimensions['J'].width = 25  # 错误输出
    
    # 生成文件名
    filename = f"execution_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 返回Excel文件
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{execution_id}", response_model=ExecutionDetail)
def get_execution_detail(
    execution_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取执行详情
    
    Args:
        execution_id: 执行ID
        db: 数据库会话
        current_user: 当前用户
    
    Returns:
        执行详情
    
    Raises:
        HTTPException: 执行记录不存在时抛出404错误
    """
    execution = db.query(ExecutionHistory).filter(ExecutionHistory.id == execution_id).first()
    if not execution:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Execution not found"
        )
    
    data = execution.to_dict()
    script = db.query(Script).filter(Script.id == execution.script_id).first()
    server = db.query(Server).filter(Server.id == execution.server_id).first()
    data['script_name'] = script.name if script else "未知"
    data['server_hostname'] = server.hostname if server else "未知"
    data['server_ip'] = server.ip_address if server else "未知"
    
    return ExecutionDetail(**data)
